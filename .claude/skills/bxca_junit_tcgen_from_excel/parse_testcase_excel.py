#!/usr/bin/env python3
"""
테스트케이스 Excel 파일을 파싱하여 JSON으로 출력하는 스크립트.
bxca_testcase_gen 스킬이 생성한 Excel 포맷을 읽습니다.

사용법: python parse_testcase_excel.py <excel_path>

출력: JSON (stdout, UTF-8)
- 성공 시: {"status": "ok", "test_cases": [...], "summary": {...}}
- 실패 시: {"status": "error", "message": "...", "expected_format": {...}}
"""

import io
import json
import sys
import os

# [#8] Windows 환경 한글 깨짐 방지: stdout을 UTF-8로 강제 설정
if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")


def ensure_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl
        return openpyxl


# [#10] 헤더 검증용 키워드 — 정확 일치 대신 핵심 키워드 포함 여부로 판별
HEADER_KEYWORDS = [
    (0, ["no"]),
    (1, ["테스트케이스", "소스파일"]),
    (2, ["테스트케이스명"]),
    (3, ["테스트내용"]),
    (4, ["테스트조건"]),
    (5, ["예상결과"]),
]

PRIORITY_KEYWORDS = ["중요도"]
TEST_RESULT_KEYWORDS = ["테스트결과"]

EXPECTED_FORMAT_DESC = {
    "description": "bxca_testcase_gen 스킬이 생성한 테스트케이스 Excel 파일",
    "required_sheet": "'테스트케이스'를 포함하는 시트명 (예: 테스트케이스_자동생성)",
    "columns_8col": ["No", "테스트케이스ID(소스파일명)", "테스트케이스명", "테스트내용",
                     "테스트조건", "예상결과", "중요도", "테스트결과"],
    "columns_7col": ["No", "테스트케이스ID(소스파일명)", "테스트케이스명", "테스트내용",
                     "테스트조건", "예상결과", "테스트결과"],
    "note": "첫 번째 행이 헤더, 두 번째 행부터 데이터. '요약' 시트는 선택사항.",
}


def normalize_header(val):
    """셀 값을 정규화 (소문자, 공백/줄바꿈 통합)."""
    if val is None:
        return ""
    return str(val).strip().replace("\r\n", " ").replace("\r", " ").replace("\n", " ").lower()


def header_contains_keywords(header_val, keywords):
    """[#10] 헤더 값이 모든 키워드를 포함하는지 유연하게 검사."""
    normalized = normalize_header(header_val)
    return all(kw.lower() in normalized for kw in keywords)


def parse_excel(excel_path):
    openpyxl = ensure_openpyxl()

    if not os.path.exists(excel_path):
        return {
            "status": "error",
            "message": f"파일을 찾을 수 없습니다: {excel_path}",
            "expected_format": EXPECTED_FORMAT_DESC,
        }

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        return {
            "status": "error",
            "message": f"Excel 파일을 열 수 없습니다: {e}",
            "expected_format": EXPECTED_FORMAT_DESC,
        }

    # [#11] try-finally로 wb.close() 보장
    try:
        return _parse_workbook(wb)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _parse_workbook(wb):
    """workbook 파싱 (wb.close()는 호출자가 담당)."""

    # 시트 찾기
    target_sheet = None
    for name in wb.sheetnames:
        if "테스트케이스" in name and "요약" not in name:
            target_sheet = wb[name]
            break

    if target_sheet is None:
        return {
            "status": "error",
            "message": f"'테스트케이스' 관련 시트를 찾을 수 없습니다. 시트 목록: {wb.sheetnames}",
            "expected_format": EXPECTED_FORMAT_DESC,
        }

    # 헤더 읽기
    rows = list(target_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    if not rows:
        return {
            "status": "error",
            "message": "시트에 데이터가 없습니다.",
            "expected_format": EXPECTED_FORMAT_DESC,
        }

    raw_headers = [h for h in rows[0] if h is not None]

    if len(raw_headers) < 6:
        return {
            "status": "error",
            "message": f"컬럼 수가 부족합니다. 최소 6열(핵심 필드) 필요, 현재 {len(raw_headers)}열.",
            "expected_format": EXPECTED_FORMAT_DESC,
        }

    # [#10] 핵심 헤더 키워드 매칭 (유연한 검증)
    mismatches = []
    for col_idx, keywords in HEADER_KEYWORDS:
        if col_idx >= len(raw_headers):
            mismatches.append(f"  열 {col_idx + 1}: 누락 (필요: {', '.join(keywords)})")
        elif not header_contains_keywords(raw_headers[col_idx], keywords):
            mismatches.append(
                f"  열 {col_idx + 1}: '{raw_headers[col_idx]}' (필요 키워드: {', '.join(keywords)})"
            )

    if mismatches:
        return {
            "status": "error",
            "message": "헤더가 예상 포맷과 다릅니다.\n"
                       f"실제 헤더: {[str(h) for h in raw_headers[:8]]}\n"
                       "불일치 항목:\n" + "\n".join(mismatches),
            "expected_format": EXPECTED_FORMAT_DESC,
        }

    # 중요도/테스트결과 열 위치 자동 감지
    has_priority = False
    priority_col = None
    test_result_col = None

    for i in range(6, min(len(raw_headers), 10)):
        if header_contains_keywords(raw_headers[i], PRIORITY_KEYWORDS):
            has_priority = True
            priority_col = i
        elif header_contains_keywords(raw_headers[i], TEST_RESULT_KEYWORDS):
            test_result_col = i

    # 데이터 파싱
    test_cases = []
    for row in target_sheet.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        # 소스파일명과 테스트케이스명 모두 비어있으면 스킵
        if not any(row[1:3]):
            continue

        tc = {
            "no": row[0],
            "source_file": str(row[1] or "").strip(),
            "test_name": str(row[2] or "").strip(),
            "test_content": str(row[3] or "").strip() if len(row) > 3 else "",
            "test_condition": str(row[4] or "").strip() if len(row) > 4 else "",
            "expected_result": str(row[5] or "").strip() if len(row) > 5 else "",
        }

        if has_priority and priority_col is not None and len(row) > priority_col:
            tc["priority"] = str(row[priority_col] or "중").strip()
        else:
            tc["priority"] = "중"

        if test_result_col is not None and len(row) > test_result_col:
            tc["test_result"] = str(row[test_result_col] or "").strip()
        else:
            tc["test_result"] = ""

        if tc["source_file"]:
            test_cases.append(tc)

    if not test_cases:
        return {
            "status": "error",
            "message": "테스트케이스 데이터가 없습니다 (헤더만 있고 데이터 행이 없음).",
            "expected_format": EXPECTED_FORMAT_DESC,
        }

    # 요약 정보
    from collections import Counter
    file_counts = Counter(tc["source_file"] for tc in test_cases)
    priority_counts = Counter(tc["priority"] for tc in test_cases)

    return {
        "status": "ok",
        "sheet_name": target_sheet.title,
        "has_priority": has_priority,
        "total_count": len(test_cases),
        "test_cases": test_cases,
        "summary": {
            "by_file": dict(file_counts),
            "by_priority": dict(priority_counts),
        },
    }


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(json.dumps({
            "status": "error",
            "message": f"Usage: {sys.argv[0]} <excel_path>",
        }, ensure_ascii=False))
        sys.exit(1)

    result = parse_excel(sys.argv[1])
    print(json.dumps(result, ensure_ascii=False, indent=2))
    sys.exit(0 if result["status"] == "ok" else 1)
