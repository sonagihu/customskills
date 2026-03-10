#!/usr/bin/env python3
"""
테스트케이스 Excel 생성 스크립트
사용법: python generate_excel.py <output_path> <json_file>

JSON 입력 형식:
{
  "sheet_name": "테스트케이스_자동생성",
  "target_directory": "c:\\DevHome\\repository\\spapp",
  "test_cases": [
    {
      "id": "소스파일명",
      "name": "테스트케이스명",
      "content": "테스트내용",
      "condition": "테스트조건",
      "expected": "예상결과",
      "priority": "상"
    }
  ]
}
"""

import json
import sys
import os
from datetime import datetime
from collections import Counter


def ensure_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl
        return openpyxl


def resolve_output_path(base_path):
    """동일 파일이 존재하면 _v2, _v3 ... 순서로 버전 suffix를 붙인다."""
    if not os.path.exists(base_path):
        return base_path

    name, ext = os.path.splitext(base_path)
    version = 2
    while True:
        versioned = f"{name}_v{version}{ext}"
        if not os.path.exists(versioned):
            return versioned
        version += 1


def generate(output_path: str, data: dict):
    openpyxl = ensure_openpyxl()
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 파일 버전 처리
    output_path = resolve_output_path(output_path)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    test_cases = data.get("test_cases", [])
    target_dir = data.get("target_directory", "N/A")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # === 공통 스타일 ===
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    cell_font = Font(name="Arial", size=10)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="top")
    label_font = Font(name="Arial", bold=True, size=10)
    value_font = Font(name="Arial", size=10)
    title_font = Font(name="Arial", bold=True, size=14, color="333333")

    # 중요도별 배경색
    priority_fills = {
        "상": PatternFill("solid", fgColor="FCE4EC"),  # 연한 빨강
        "중": PatternFill("solid", fgColor="FFF9C4"),  # 연한 노랑
        "하": None,
    }

    # =============================================
    # 시트 1: 요약
    # =============================================
    ws_summary = wb.create_sheet(title="요약")

    # 타이틀
    ws_summary.cell(row=1, column=1, value="테스트케이스 요약").font = title_font
    ws_summary.merge_cells("A1:D1")
    ws_summary.row_dimensions[1].height = 30

    # 생성 정보
    ws_summary.cell(row=3, column=1, value="생성 일시").font = label_font
    ws_summary.cell(row=3, column=2, value=now).font = value_font
    ws_summary.cell(row=4, column=1, value="작업 대상").font = label_font
    ws_summary.cell(row=4, column=2, value=target_dir).font = value_font
    ws_summary.cell(row=5, column=1, value="총 테스트케이스").font = label_font
    ws_summary.cell(row=5, column=2, value=len(test_cases)).font = value_font

    # 파일별 건수 집계
    file_counts = Counter(tc.get("id", "") for tc in test_cases)
    ws_summary.cell(row=7, column=1, value="파일별 테스트케이스 건수").font = label_font
    summary_headers = ["소스파일명", "건수"]
    for col_idx, text in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=8, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, (file_name, count) in enumerate(file_counts.items(), 9):
        cell_name = ws_summary.cell(row=row_idx, column=1, value=file_name)
        cell_name.font = cell_font
        cell_name.border = thin_border
        cell_count = ws_summary.cell(row=row_idx, column=2, value=count)
        cell_count.font = cell_font
        cell_count.border = thin_border
        cell_count.alignment = center_alignment

    # 합계 행
    total_row = 9 + len(file_counts)
    cell_total_label = ws_summary.cell(row=total_row, column=1, value="합계")
    cell_total_label.font = Font(name="Arial", bold=True, size=10)
    cell_total_label.border = thin_border
    cell_total_value = ws_summary.cell(row=total_row, column=2, value=len(test_cases))
    cell_total_value.font = Font(name="Arial", bold=True, size=10)
    cell_total_value.border = thin_border
    cell_total_value.alignment = center_alignment

    # 중요도별 건수 집계
    priority_counts = Counter(tc.get("priority", "중") for tc in test_cases)
    priority_start = total_row + 2
    ws_summary.cell(row=priority_start, column=1, value="중요도별 건수").font = label_font
    priority_headers = ["중요도", "건수"]
    for col_idx, text in enumerate(priority_headers, 1):
        cell = ws_summary.cell(row=priority_start + 1, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_offset, priority in enumerate(["상", "중", "하"]):
        row = priority_start + 2 + row_offset
        cell_p = ws_summary.cell(row=row, column=1, value=priority)
        cell_p.font = cell_font
        cell_p.border = thin_border
        cell_p.alignment = center_alignment
        fill = priority_fills.get(priority)
        if fill:
            cell_p.fill = fill
        cell_c = ws_summary.cell(row=row, column=2, value=priority_counts.get(priority, 0))
        cell_c.font = cell_font
        cell_c.border = thin_border
        cell_c.alignment = center_alignment

    # 요약 시트 열 너비
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 20

    # =============================================
    # 시트 2: 테스트케이스_자동생성
    # =============================================
    sheet_name = data.get("sheet_name", "테스트케이스_자동생성")
    ws = wb.create_sheet(title=sheet_name)

    # 헤더 정의 (중요도 필드 추가)
    headers = [
        ("No", 6),
        ("테스트케이스ID\n(소스파일명)", 30),
        ("테스트케이스명", 35),
        ("테스트내용", 50),
        ("테스트조건", 40),
        ("예상결과", 40),
        ("중요도", 10),
        ("테스트결과", 15),
    ]

    for col_idx, (header_text, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 35

    # 데이터 작성
    for row_idx, tc in enumerate(test_cases, start=2):
        priority = tc.get("priority", "중")
        values = [
            row_idx - 1,
            tc.get("id", ""),
            tc.get("name", ""),
            tc.get("content", ""),
            tc.get("condition", ""),
            tc.get("expected", ""),
            priority,
            "",  # 테스트결과는 공란
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border
            if col_idx in (1, 7, 8):  # No, 중요도, 테스트결과
                cell.alignment = center_alignment
            else:
                cell.alignment = cell_alignment

            # 중요도 열 배경색
            if col_idx == 7:
                fill = priority_fills.get(priority)
                if fill:
                    cell.fill = fill

    # 자동 필터
    if test_cases:
        last_row = len(test_cases) + 1
        ws.auto_filter.ref = f"A1:H{last_row}"

    # 틀 고정
    ws.freeze_panes = "A2"

    # 저장
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    print(f"Generated: {output_path} ({len(test_cases)} test cases)")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(f"Usage: {sys.argv[0]} <output_xlsx_path> <json_file>", file=sys.stderr)
        sys.exit(1)

    output_path = sys.argv[1]
    json_file = sys.argv[2]

    with open(json_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    generate(output_path, data)
